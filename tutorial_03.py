from openpyxl.workbook import Workbook
# from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

#  openpyxl ia a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files

print(f"<Start of tutorial_03.py>\n")

# Lesson 15 - Create A Spreadsheet Workbook
print(f"Lesson 15 - Create A Spreadsheet Workbook\n")

# Create workbook object
wb = Workbook()

# Create an active worksheet
ws = wb.active

# Create worksheet title
ws.title = "Names and Colours"

# Lesson 16 - Add Data To New Spreadsheet Workbook
print(f"Lesson 16 - Add Data To New Spreadsheet Workbook\n")

# Create Python List of Name and colours
names = ["Westham", "Arsenal", "Chelsea", "Man Utd", "Newcastle", "Wolves", "Leeds"]
colours = ["Claret", "Yellow", "Blue", "Red", "Black", "Orange", "White"]

# Loop through names and add each name
ws['A1'] = "Team"  # Set Column Header Title
ws['A1'].font = Font(bold=True)
ws['B1'] = "Colour"  # Set Column Header Title
ws['B1'].font = Font(bold=True)

starting_row = 2  # Set Starting Row
for name in names:  # Loop through names to be added to column 1
    ws.cell(row=starting_row, column=1).value = name  # Enter name into cell
    starting_row += 1  # increment and step into next name

starting_row = 2  # Reset Starting Row back to 2
for colour in colours:  # Loop through colours to be added to column 2
    ws.cell(row=starting_row, column=2).value = colour  # Enter name into cell
    starting_row += 1  # increment and step into next name

# Save Spreadsheet
# wb.save('tutorial_03.xlsx')  #Save at end of document for version no

# Lesson 17 - Use Excel Formulas With Python
print(f"Lesson 17 - Use Excel Formulas With Python\n")

# Add column that holds a number
ws['C1'] = "League Position 2023"
ws['C1'].font = Font(bold=True)
positions = [15, 2, 12, 3, 4, 13, 19]

starting_row = 2  # Reset Starting Row back to 2
for position in positions:  # Loop through colours to be added to column 3
    ws.cell(row=starting_row, column=3).value = position  # Enter name into cell
    starting_row += 1  # increment and step into next name

# Create formula
ws['C9'] = "=AVERAGE(C2:C8)" #  Adds the average of range to this cell

# Save Spreadsheet
# wb.save('tutorial_03.xlsx')  See new method below

# Updated save method to allow to close excel and save update
while True:
    try:
        wb.save('tutorial_03.xlsx')
    except PermissionError:
        input("Please close 'Excel' and press Enter")
    else:
        print("File saved...")
        break

# Lesson 18 - Change Cell Font, Size, Color, Boldness, Italics
print(f"Lesson 18 - Change Cell Font, Size, Color, Boldness, Italics\n")

# We need to import font in opening statements

cell = ws['A1']  # Select cell

# Make changes to just this cell
cell.font = Font(
    size=20,
    bold=True,
    italic=False,
    color="aa0011"

)

# Lesson 19 - Add Borders To Cells in a Spreadsheet
print(f"Lesson 19 - Add Borders To Cells in a Spreadsheet\n")

# We need to import border and side in opening statements

# Define a border

my_border = Side(style="thick", color="000000")  # Set style
C9 = ws['C9']  # Set cell as variable

# Now Set Border
C9.border = Border(
    left=my_border,
    right=my_border,
    top=my_border,
    bottom=my_border)


# Save Spreadsheet (quick save)
wb.save('tutorial_03.xlsx')

print('File Saved to tutorial_03.xlsx...')

print(f'\n{"<End of tutorial_03.py>"}')  # Add newline to start line


