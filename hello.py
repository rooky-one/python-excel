from openpyxl.workbook import Workbook
from openpyxl import load_workbook

print(f"<Start of hello.py>\n")
# Create workbook object
# wb = Workbook()

# load existing spreadsheet
wb = load_workbook('hello.xlsx')

# Create an active worksheet
ws = wb.active

# Use a variable to hold spreadsheet call value
name = ws['A2'].value
colour = ws['B2'].value

# Print something from our spreadsheet
print(ws['A2'].value)
print(ws['B2'].value)
print(f'{name}: {colour}\n')

# Grab a whole column
column_a = ws['A']  # Returns immutable tuple
column_b = ws['B']  # Returns immutable tuple

print(f'{column_a} \n')  # Prints the immutable tuple list
print(f'{column_b} \n')  # Prints the immutable tuple list

# Loop through the tuple
for spreadsheet_cells in column_a:
    print(spreadsheet_cells.value)  # Print tuple contents as list

print('')

# Loop through the tuple
for spreadsheet_cells in column_b:
    print(spreadsheet_cells.value)  # Print tuple contents as list

print('')  # Add blank line

# Lesson 9

# Grab a Range
cell_range = ws['A2':'A11']  # Just names exclude the column name
print(range)  # See note...
# Note the first 2 characters are (( this indicates a tuples in a tuple
# This means to extract the cell data we need to loop through at least twice

# Create one loop
for cells_in_range in cell_range:
    print(cells_in_range)  # Prints the cell ID not data held within

print('')  # Add blank line

# Create two loops
for cells_in_range in cell_range:  # Set range
    for x in cells_in_range:  # Loop through entire range
        print(x.value)  # .value pulls out cell data

print('')  # Add blank line

# Grab a 2 Column in Range
cell_range = ws['A2':'B11']  # Exclude the column names

# Create two loops (gives column A then Column B in list)
for cells_in_range in cell_range:  # Set range
    for x in cells_in_range:  # Loop through entire range
        print(x.value)  # .value pulls out cell data

print(f'\n{"<End of hello.py>"}')  # Add newline to start line
