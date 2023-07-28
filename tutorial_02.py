from openpyxl.workbook import Workbook
from openpyxl import load_workbook
#  openpyxl ia a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files

print(f"<Start of tutorial_02.py>\n")
# Create workbook object
# wb = Workbook()

# load existing spreadsheet
wb = load_workbook('tutorial_02.xlsx')

# Create an active worksheet
ws = wb.active

# Lesson 12 - Change Cells and Save Spreadsheet (1st Method)
print(f"Lesson 12 - Change Cells and Save Spreadsheet (1st Method)\n")

# Change a cell

ws["A11"] = "Judas"

# Save the spreadsheet (two options)
# wb.save('C:\code\python-excel\tutorial_02.xlsx')  # Absolute path
wb.save('tutorial_02.xlsx')  # Relative save, note save under new file name

print("File was saved...")
print("")

# Lesson 13 - Add Cells To A Spreadsheet (Second Method)
print(f"Lesson 13 - Add Cells To A Spreadsheet (Second Method)\n")

# Change a many cells

starting_row = 12
ws.cell(row=11, column=1).value = "Bartholomew"
ws.cell(row=11, column=2).value = "Black"

wb.save('tutorial_02.xlsx')

print("")
print("File was saved...")
print("")

# Lesson 14 - Loop Through a Spreadsheet and Add Names
print(f"Lesson 14 - Loop Through a Spreadsheet and Add Names\n")

# Create Python List of Names
names = ["Jan", "Feb", "Mar"]

# Change many rows
starting_row = 13

for name in names:
    ws.cell(starting_row, column=1).value = name  # Loop through names
    starting_row += 1  # increment counter

wb.save('tutorial_02.xlsx')

print("")
print("File was saved...")

print(f'\n{"<End of tutorial_02.py>"}')  # Add newline to start line
