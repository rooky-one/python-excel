from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.chart import BarChart, BarChart3D, Reference

print(f"<Start of tutorial_05.py>\n")
# Create workbook object
# wb = Workbook()

# load existing spreadsheet
wb = load_workbook('tutorial_05.xlsx')

# Create an active worksheet
ws = wb.active

# Lesson 21 - Create a Bar Chart In A Spreadsheet With Python
print(f"Lesson 21 - Create a Bar Chart In A Spreadsheet With Python\n")

# Determine Type of Chart
chart = BarChart()

# Labels and Data
labels = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=11)
data = Reference(ws, min_col=3, min_row=1, max_row=11)

# Tie it is together
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

# Add title
chart.title = "Employees Salaries"

# Place chart onto spreadsheet
ws.add_chart(chart, "E2")

# Save Spreadsheet (quick save)
wb.save('tutorial_05.xlsx')

print('File Saved to tutorial_05.xlsx...')

print(f'\n{"<End of tutorial_05.py>"}')  # Add newline to start line
