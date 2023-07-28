from openpyxl.workbook import Workbook
from openpyxl import load_workbook

print(f"<Start of tutorial_01.py>\n")
# Create workbook object
# wb = Workbook()

# load existing spreadsheet
wb = load_workbook('tutorial_01.xlsx')

# Create an active worksheet
ws = wb.active

# Use a variable to hold spreadsheet call value
name = ws['A2'].value
colour = ws['B2'].value

# Print something from our spreadsheet
print(ws['A2'].value)
print(ws['B2'].value)
print(f'{name}: {colour}\n')

# Grab a whole column
column_a = ws['A']  # Returns immutable tuple
column_b = ws['B']  # Returns immutable tuple

print(f'{column_a} \n')  # Prints the immutable tuple list
print(f'{column_b} \n')  # Prints the immutable tuple list

# Loop through the tuple
for spreadsheet_cells in column_a:
    print(spreadsheet_cells.value)  # Print tuple contents as list

print('')

# Loop through the tuple
for spreadsheet_cells in column_b:
    print(spreadsheet_cells.value)  # Print tuple contents as list

print('')
print('Lesson 9')  # Add blank line
print('')

# Lesson 9

# Grab a Range
cell_range = ws['A2':'A11']  # Just names exclude the column name
print(range)  # See note...
# Note the first 2 characters are (( this indicates a tuples in a tuple
# This means to extract the cell data we need to loop through at least twice

# Create one loop
for cells_in_range in cell_range:
    print(cells_in_range)  # Prints the cell ID not data held within

print('')  # Add blank line

# Create two loops
for cells_in_range in cell_range:  # Set range
    for x in cells_in_range:  # Loop through entire range
        print(x.value)  # .value pulls out cell data

print('')  # Add blank line

# Grab a 2 Column in Range
cell_range = ws['A2':'B11']  # Exclude the column names

# Create two loops (gives column A then Column B in list)
for cells_in_range in cell_range:  # Set range
    for x in cells_in_range:  # Loop through entire range
        print(x.value)  # .value pulls out cell data

print('')
print('Lesson 10')
print('')

# Lesson 10 Iterate through rows of a spreadsheet
for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=2, values_only=True):  # 2 starts from row 2, ends row 7
    #  value_only save us having to use the .value argument
    print(row)

print('')


for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=2, values_only=True):  # 2 starts from row 2, ends row 7
    #  value_only save us having to use the .value argument
    #print(row)
    for cell in row:
        print(cell)

print('')
print('Lesson 11')
print('')

# Lesson 11 Iterate through columns of a spreadsheet
# Give same results as iterate by row, just a different way
for cols in ws.iter_cols(min_row=2, max_row=7, min_col=1, max_col=2, values_only=True):  # 2 starts from row 2, ends row 7
    #  value_only save us having to use the .value argument
    print(cols)

print('')

for cols in ws.iter_cols(min_row=2, max_row=7, min_col=1, max_col=2, values_only=True):  # 2 starts from row 2, ends row 7
    #  value_only save us having to use the .value argument
    #print(cols)
    for cell in cols:
        print(cell)

print(f'\n{"<End of tutorial_01.py>"}')  # Add newline to start line
